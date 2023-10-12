#!/bin/python

###TODO!!!
# Fungsi input data
## Fungsi selective insert (SQL)
# Uji fungsi cari lokasi via text

### Importing necessary libraries
import asyncio
import csv
from datetime import datetime, timedelta
from functools import wraps, partial
from geopy.geocoders import Nominatim
import haversine as hs
import io
# import logging
import jwt # pip install pyjwt
from math import ceil, radians, sin, cos, sqrt, atan2
# import MySQLdb # pip install mysqlclient
import os
import openpyxl
import pymysql
import signal
import secrets
from string import ascii_letters, digits
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.constants import ParseMode
import telegram.error
from telegram.ext import Application, CommandHandler, ConversationHandler, MessageHandler, filters, ContextTypes, CallbackQueryHandler
import time

# Instantiate global variables
SQRTo2 = sqrt(2)
db_connection = None
jwt_token_keys = {}

def create_mysql_connection(passwd: str):
    print('Trying to establish connection to the database...', flush = True)

    try:
        conn = pymysql.connect(
            host = os.environ['MYSQL_HOSTNAME'],
            port = int(os.environ['MYSQL_PORT']),
            user = os.environ['TELEBOT_USER'],
            password=passwd,
            database = os.environ['MYSQL_DATABASE']
        )
        return conn
    except pymysql.Error as e:
        print(f"Error connecting to the database: {e}")
        return None

def close_db_connection(connection):
    if connection:
        connection.close()
        print('Database connection closed', flush = True)

# Main Function
def main():
    # Initializing Configuration
    print('Initializing configuration...', flush = True)
    BOT_TOKEN = os.environ['BOT_TOKEN']

    app = Application.builder().token(BOT_TOKEN).build()

    telebot_passwd = generate_password()

    with open(r'/init_status.txt', 'r') as file:
        init_status = file.read()[:-1]

    if init_status == 'Basis data belum terinisialisasi.':
        filename = r'/docker-entrypoint-initdb.d/init.sql'

    elif init_status == 'Basis data sudah terinisialisasi.':
        filename = r'/always-initdb.d/init.sql'

    with open(filename, 'r') as file:
        data = file.read()

    with open(filename, 'w') as file:
        data = data.replace('__TELEBOT_PASSWORD__', telebot_passwd)
        file.write(data)

    print(f'Mysql user password: {telebot_passwd}', flush  = True)

    ### Conversation Handler Menu Input
    input_data_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(menu_input, pattern='^opsi_input_data$')],
        states={
            1: [CallbackQueryHandler(input_data, pattern='^input_')],
            2: [MessageHandler(filters.Document.FileExtension('xlsx'), input_satuan_xlsx), MessageHandler(filters.Document.FileExtension('csv'), input_satuan_csv)],
            3: [MessageHandler(filters.Document.FileExtension('xlsx'), input_pangkas_xlsx), MessageHandler(filters.Document.FileExtension('csv'), input_pangkas_csv)]
        },
        fallbacks=[CommandHandler('batal', batal)]
    )

    ### Conversation Handler Menu Tambah User
    tambah_user_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(tambah_user, pattern='^opsi_tambah_user$')],
        states={
            1: [CallbackQueryHandler(input_tambah_user, pattern='^tambah_user_')],
            2: [MessageHandler(filters.TEXT & (~ filters.COMMAND), proses_tambah_user)],
            3: [MessageHandler(filters.TEXT & (~ filters.COMMAND), proses_tambah_admin)]
        },
        fallbacks=[CommandHandler('batal', batal)]
    )

    ### Conversation Handler Menu Hapus User
    hapus_user_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(hapus_user, pattern='^opsi_hapus_user$')],
        states={
            1: [CallbackQueryHandler(konfirmasi_hapus_user, pattern='^hapus_user_')],
            2: [CallbackQueryHandler(proses_hapus_user, pattern='^konfirmasi_hapus_user_')]
        },
        fallbacks=[CommandHandler('batal', batal)]
    )

    ### Conversation Handler Menu Tambah Contact Person
    tambah_cp_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(tambah_cp, pattern='^opsi_tambah_cp$')],
        states={
            1: [MessageHandler(filters.TEXT & (~ filters.COMMAND), input_tambah_cp)],
            2: [MessageHandler(filters.TEXT & (~ filters.COMMAND), proses_tambah_cp)]
        },
        fallbacks=[CommandHandler('batal', batal)]
    )

    ### Conversation Handler Menu Hapus Contact Person
    hapus_cp_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(hapus_cp, pattern='^opsi_hapus_cp$')],
        states={
            1: [CallbackQueryHandler(konfirmasi_hapus_cp, pattern='^hapus_cp_')],
            2: [CallbackQueryHandler(proses_hapus_cp, pattern='^konfirmasi_hapus_cp_')]
        },
        fallbacks=[CommandHandler('batal', batal)]
    )

    ### Conversation Handler Menu Peroleh Lokasi
    peroleh_lokasi_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(peroleh_lokasi, pattern='^opsi_peroleh_lokasi$')],
        states={
            1: [CallbackQueryHandler(proses_peroleh_lokasi_button, pattern='^item_'), MessageHandler(filters.TEXT & (~ filters.COMMAND), proses_peroleh_lokasi_text)]
        },
        fallbacks=[CommandHandler('batal', batal)]
    )

    ### Conversation Handler Menu Peroleh Daftar
    peroleh_nama_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(peroleh_nama_1, pattern='^opsi_peroleh_nama$')],
        states={
            1: [MessageHandler(filters.LOCATION, peroleh_nama_2)],
            2: [MessageHandler(filters.TEXT & (~ filters.COMMAND), proses_peroleh_nama)]
        },
        fallbacks=[CommandHandler('batal', batal)]
    )

    ### Conversation Handler Menu Peroleh File
    peroleh_berkas_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(peroleh_berkas, pattern='^opsi_peroleh_berkas$')],
        states={
            1: [MessageHandler(filters.TEXT & (~ filters.COMMAND), proses_peroleh_berkas)]
        },
        fallbacks=[CommandHandler('batal', batal)]
    )

    ### Conversation Handler Token
    peroleh_token_handler = ConversationHandler(
        entry_points=[CommandHandler('peroleh_token', peroleh_token)],
        states={
            1: [MessageHandler(filters.TEXT & (~ filters.COMMAND), peroleh_token_process)]
        },
        fallbacks=[CommandHandler('batal', batal)]
    )

    # Handlers
    app.add_handler(CommandHandler('start', main_menu))
    app.add_handler(input_data_handler)
    app.add_handler(tambah_user_handler)
    app.add_handler(hapus_user_handler)
    app.add_handler(tambah_cp_handler)
    app.add_handler(hapus_cp_handler)
    app.add_handler(peroleh_lokasi_handler)
    app.add_handler(CommandHandler('peroleh_lokasi', peroleh_lokasi_func))
    app.add_handler(peroleh_nama_handler)
    app.add_handler(peroleh_berkas_handler)
    app.add_handler(peroleh_token_handler)
    app.add_handler(CommandHandler('peroleh_username', peroleh_username))
    app.add_handler(CommandHandler('bantuan', kirim_bantuan))
    app.add_handler(CallbackQueryHandler(kirim_bantuan, pattern='^opsi_bantuan$'))

    global db_connection
    while db_connection == None:
        time.sleep(5)
        db_connection = create_mysql_connection(telebot_passwd)

    print('Connection to the database has been established.', flush = True)

    while True:
        try:
            print('Polling...', flush = True)
            app.run_polling(poll_interval=4)
        except telegram.error.TimedOut as e:
            print(e, flush = True)
            app.stop_running()
            print('Polling temporarily stopped.', flush = True)
            time.sleep(10)
            continue

def exclude_commands_filter(update):
    return not update.message.text.startswith('/')

def generate_password(pass_len: int=32):
  alphabet = ascii_letters + digits + '~!@#$^&*()-+={}[]/<>,.?:| '
  unescaped = ''.join(secrets.choice(alphabet) for i in range(pass_len))
  return unescaped

### Middleware functions for user authentication
def authenticate_user(func):
    @wraps(func)
    async def wrapper(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False], *args, **kwargs):
        # Check if the user is authenticated
        auth_status_res = is_authenticated(update.effective_user.id)
        if auth_status_res[0]:
            # User is authenticated, execute the command handler
            return await func(update, context, auth_status_res, *args, **kwargs)
        else:
            # User is not authenticated.
            await retry_on_error(context.bot.send_message, chat_id=update.effective_chat.id, text='Access denied.')
    return wrapper

def authenticate_admin(func):
    @wraps(func)
    async def wrapper(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False], *args, **kwargs):
        # Check if the user is authenticated and is an admin
        auth_status_res = is_authenticated(update.effective_user.id)
        if auth_status_res[0] and auth_status_res[1]:
            # Admin is authenticated, execute the command handler
            return await func(update, context, auth_status_res, *args, **kwargs)
        else:
            # User is not authenticated or not an admin.
            await retry_on_error(context.bot.send_message, chat_id=update.effective_chat.id, text='Access denied.')
    return wrapper

# Function to handle network errors
async def retry_on_error(func, wait=0.15, retry=-1, *args, **kwargs):
    i = 0
    while True:
        try:
            return await func(*args, **kwargs)
            break
        except telegram.error.NetworkError:
            time.sleep(wait)
            if i == retry:
                break
            i += 1
            print(f'Network Error. Retrying...{i}', flush = True)

# Function to check user access in the database
def is_authenticated(user_id):
    query_admin = 'SELECT `is_admin` FROM `allowed_users` WHERE `username` = %s'

    global db_connection
    with db_connection.cursor() as cursor:
        cursor.execute(query_admin, (user_id,))
        result = cursor.fetchone()

    if result:
        if result[0]:
            return [True, True]  # Return True if user is an admin with full access
        else:
            return [True, False]  # Return True if user is an admin with full access
    return [False, False]

def check_conv_status(func):
    @wraps(func)
    async def wrapper(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False], *args, **kwargs):
        if context.chat_data.get('in_conversation'):
            try:
                await retry_on_error(update.message.reply_text, text='Mohon akhiri percakapan terlebih dahulu dengan menjalankan fungsi /batal.')
            except AttributeError:
                await retry_on_error(context.bot.send_message, chat_id=query.message.chat_id, text='Mohon akhiri percakapan terlebih dahulu dengan menjalankan fungsi /batal.', reply_markup=InlineKeyboardMarkup(keyboard))
        else:
            # User is not authenticated or not an admin.
            return await func(update, context, auth_status, *args, **kwargs)
    return wrapper

# Retrieve data from MySQL
def peroleh_nama_site():
    query = 'SELECT `Site_ID_Tenant` FROM `site_data`'

    global db_connection
    with db_connection.cursor() as cursor:
        cursor.execute(query)
        site_ids = [Site_ID_Tenant[0] for Site_ID_Tenant in cursor]
    return site_ids

def cek_nama_site(site_id):
    query = 'SELECT 1 FROM `site_data` WHERE `Site_ID_Tenant` = %s'

    global db_connection
    with db_connection.cursor() as cursor:
        cursor.execute(query, (site_id))
        exists = cursor.fetchone()

    return exists

# Retrieve Sites Datas from MySQL
def peroleh_data_site(site_id):
    query = 'SELECT `Site_ID_Tenant`, `Tenant`, `Alamat`, `Latitude`, `Longitude` FROM `site_data` WHERE `Site_ID_Tenant` = %s'

    global db_connection
    with db_connection.cursor() as cursor:
        cursor.execute(query, (site_id))
        site_data = cursor.fetchone()

    return site_data

# Retrieve datas using radius and location
def peroleh_dari_radius(latitude, longitude, radius):
    coordinates = (latitude, longitude)
    global SQRTo2
    sqrt_rds = radius * SQRTo2
    max_lat = hs.inverse_haversine(coordinates, sqrt_rds, hs.Direction.NORTH, unit=hs.Unit.METERS)[0]
    max_lon = hs.inverse_haversine(coordinates, sqrt_rds, hs.Direction.EAST, unit=hs.Unit.METERS)[1]
    min_lat = hs.inverse_haversine(coordinates, sqrt_rds, hs.Direction.SOUTH, unit=hs.Unit.METERS)[0]
    min_lon = hs.inverse_haversine(coordinates, sqrt_rds, hs.Direction.WEST, unit=hs.Unit.METERS)[1]
    query = 'SELECT `Site_ID_Tenant`, `Tenant`, `Alamat`, `Latitude`, `Longitude` FROM `site_data` WHERE `Latitude` BETWEEN %s AND %s AND `Longitude` BETWEEN %s AND %s'

    global db_connection
    with db_connection.cursor() as cursor:
        cursor.execute(query, (min_lat, max_lat, min_lon, max_lon))
        square_res = cursor.fetchall()
    results = []

    for item in square_res:
        dist = hs.haversine((latitude, longitude), (item[3], item[4]), unit=hs.Unit.METERS)
        if dist <= radius:
            results.append(item)
    return results

# Insert data into MySQL
def truncate_and_insert_sites(site_datas):
    truncate = 'TRUNCATE TABLE `site_data`;'
    insert = 'INSERT INTO `site_data` (`Site_ID_Tenant`, `Tenant`, `Alamat`, `Latitude`, `Longitude`) VALUES (%s, %s, %s, %s, %s);'

    global db_connection
    with db_connection.cursor() as cursor:
        cursor.execute(truncate)
        cursor.executemany(insert, [list(dictionary.values()) for dictionary in site_datas])

    db_connection.commit()

def insert_sites(site_datas):
    insert = 'INSERT INTO `site_data` (`Site_ID_Tenant`, `Tenant`, `Alamat`, `Latitude`, `Longitude`) VALUES (%s, %s, %s, %s, %s);'

    global db_connection
    with db_connection.cursor() as cursor:
        cursor.executemany(insert, [list(dictionary.values()) for dictionary in site_datas])

    db_connection.commit()

### Main Menu
@authenticate_user
@check_conv_status
async def main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    keyboard = [
        [InlineKeyboardButton('Peroleh Lokasi Item', callback_data='opsi_peroleh_lokasi')],
        [InlineKeyboardButton('Peroleh Daftar Item Terdekat', callback_data='opsi_peroleh_nama')],
        [InlineKeyboardButton('Peroleh File Site', callback_data='opsi_peroleh_berkas')],
        [InlineKeyboardButton('Bantuan', callback_data='opsi_bantuan')]
    ]

    if auth_status[1]:
        keyboard =  [
            [InlineKeyboardButton('Input Data', callback_data='opsi_input_data')],
            [InlineKeyboardButton('Tambah User', callback_data='opsi_tambah_user')],
            [InlineKeyboardButton('Hapus User', callback_data='opsi_hapus_user')],
            [InlineKeyboardButton('Tambah Contact Person', callback_data='opsi_tambah_cp')],
            [InlineKeyboardButton('Hapus Contact Person', callback_data='opsi_hapus_cp')],
        ] + keyboard

    await retry_on_error(update.message.reply_text, text='Menu Utama', reply_markup=InlineKeyboardMarkup(keyboard))

### Input Data
@authenticate_admin
@check_conv_status
async def menu_input(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    query = update.callback_query
    context.chat_data['in_conversation'] = True
    await context.bot.edit_message_reply_markup(chat_id=query.message.chat_id, message_id=query.message.message_id, reply_markup=None)

    keyboard = [
        [InlineKeyboardButton('Pangkas dan Input', callback_data='input_keseluruhan')],
        [InlineKeyboardButton('Input Satuan', callback_data='input_satuan')],
    ]

    await retry_on_error(context.bot.send_message, chat_id=query.message.chat_id, text='Pilih metode input.', reply_markup=InlineKeyboardMarkup(keyboard))
    return 1

@authenticate_admin
async def input_data(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    query = update.callback_query
    button_pressed = query.data

    await context.bot.edit_message_reply_markup(chat_id=query.message.chat_id, message_id=query.message.message_id, reply_markup=None)
    await retry_on_error(context.bot.send_message, chat_id=query.message.chat_id, text='Silahkan unggah berkas.')

    if button_pressed == 'input_satuan':
        return 2

    return 3

@authenticate_admin
async def input_satuan_xlsx(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    path = os.environ['APP_TMP_DATA'] + '/input.xlsx'
    await (await context.bot.get_file(update.message.document.file_id)).download_to_drive(custom_path=path)
    output = []
    wb =  openpyxl.load_workbook(path)
    ws = wb.active
    key = [ws.cell(1, col + 1).value for col in range(4)]
    count = 0

    for row in range(2, ws.max_row + 1):
        if cek_nama_site(ws.cell(row, 1).value):
            continue

        else:
            output.append({key[col]: ws.cell(row, col + 1).value for col in range(4)})
            output[-1]['Latitude'], output[-1]['Longitude'] = (float(value) for value in output[-1].pop('Koordinat Site').split(','))
            output[-1]['Alamat'] = cari_alamat((output[-1]['Latitude'], output[-1]['Longitude']))
            count += 1

    insert_sites(output)
    await retry_on_error(update.message.reply_text, text=f'Input satuan {count} baris berhasil.')
    return akhiri_percakapan(context)

@authenticate_admin
async def input_satuan_csv(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    file = await context.bot.get_file(update.message.document.file_id)
    buffer = io.BytesIO()
    await file.download_to_memory(out=buffer)
    file_size = len(buffer.getvalue())
    output = []
    site_list = peroleh_nama_site()
    count = 0

    with io.TextIOWrapper(io.BytesIO(buffer.getvalue())) as csv_file:
        csv_reader = csv.DictReader(csv_file)

        for row in csv_reader:
            if cek_nama_site(row['Site_ID_Tenant']):
                continue

            else:
                row['Latitude'], row['Longitude'] = (float(value) for value in row.pop('Koordinat Site').split(','))
                row['Alamat'] = cari_alamat((row['Latitude'], row['Longitude']))
                output.append(row)
                count += 1

    insert_sites(output)
    await retry_on_error(update.message.reply_text, text=f'Input satuan {count} baris berhasil.')
    return akhiri_percakapan(context)

@authenticate_admin
async def input_pangkas_xlsx(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    path = os.environ['APP_TMP_DATA'] + '/input.xlsx'
    await (await context.bot.get_file(update.message.document.file_id)).download_to_drive(custom_path=path)
    output = []
    wb =  openpyxl.load_workbook(path)
    ws = wb.active
    key = [ws.cell(1, col + 1).value for col in range(4)]
    count = 0

    for row in range(2, ws.max_row + 1):
        output.append({key[col]: ws.cell(row, col + 1).value for col in range(4)})
        output[-1]['Latitude'], output[-1]['Longitude'] = (float(value) for value in output[-1].pop('Koordinat Site').split(','))
        output[-1]['Alamat'] = cari_alamat((output[-1]['Latitude'], output[-1]['Longitude']))
        count += 1

    truncate_and_insert_sites(output)
    await retry_on_error(update.message.reply_text, text=f'Input pangkas {count} baris berhasil.')
    return akhiri_percakapan(context)

@authenticate_admin
async def input_pangkas_csv(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    file = await context.bot.get_file(update.message.document.file_id)
    buffer = io.BytesIO()
    await file.download_to_memory(out=buffer)
    file_size = len(buffer.getvalue())
    output = []
    count = 0

    with io.TextIOWrapper(io.BytesIO(buffer.getvalue())) as csv_file:
        csv_reader = csv.DictReader(csv_file)

        for row in csv_reader:
            row['Latitude'], row['Longitude'] = (float(value) for value in row.pop('Koordinat Site').split(','))
            row['Alamat'] = cari_alamat((row['Latitude'], row['Longitude']))
            output.append(row)
            count += 1

    truncate_and_insert_sites(output)
    await retry_on_error(update.message.reply_text, text=f'Input pangkas {count} baris berhasil.')
    return akhiri_percakapan(context)

def cari_alamat(koordinat):
    ongoing = True
    geolocator = Nominatim(user_agent='Geocoder')

    while ongoing:
        sleeptime = 1
        try:
            location = geolocator.reverse(koordinat)
            ongoing = False
            return location.address
        except Exception as e:
            print('Geocoder error:', e, 'Retrying...', flush = True)
            sleeptime += sleeptime<=15
            time.sleep(sleeptime)
            pass

    return None

### Tambah User
@authenticate_admin
@check_conv_status
async def tambah_user(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    query = update.callback_query
    await context.bot.edit_message_reply_markup(chat_id=query.message.chat_id, message_id=query.message.message_id, reply_markup=None)
    context.chat_data['in_conversation'] = True

    keyboard = [
        [InlineKeyboardButton('Tambahkan User Reguler', callback_data='tambah_user_reguler')],
        [InlineKeyboardButton('Tambahkan User Admin', callback_data='tambah_user_admin')],
    ]

    await retry_on_error(context.bot.send_message, chat_id=query.message.chat_id, text='Pilih jenis user.', reply_markup=InlineKeyboardMarkup(keyboard))
    return 1

@authenticate_admin
async def input_tambah_user(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    query = update.callback_query
    button_pressed = query.data
    await context.bot.edit_message_reply_markup(chat_id=query.message.chat_id, message_id=query.message.message_id, reply_markup=None)

    if button_pressed == 'tambah_user_admin':
        await retry_on_error(context.bot.send_message, chat_id=query.message.chat_id, text='Silahkan masukkan token calon user admin. Untuk membuat token, calon user admin tersangkut harus menjalankan perintah /peroleh_token dan memberikan nama lengkapnya dalam tahapan yang dijalankan.')
        return 3

    elif button_pressed == 'tambah_user_reguler':
        await retry_on_error(context.bot.send_message, chat_id=query.message.chat_id, text='Silahkan masukkan token calon user reguler. Untuk membuat token, calon user reguler tersangkut harus menjalankan perintah /peroleh_token dan memberikan nama lengkapnya dalam tahapan yang dijalankan.')

    return 2

@authenticate_admin
async def proses_tambah_user(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    token = update.message.text
    userid = jwt.get_unverified_header(token)['user_id']
    global jwt_token_keys
    key = jwt_token_keys.pop(userid)

    try:
        decoded_token = jwt.decode(token, key, algorithms=['HS512'])
    except jwt.ExpiredSignatureError:
        await retry_on_error(update.message.reply_text, text='Token telah kadaluarsa. Silahkan membuat token baru dan masukkan kembali token apabila token sudah dibuat kembali oleh calon user reguler tersangkut atau batalkan proses tambah user reguler dengan menjalankan perintah /batal.')
        return 2

    nama = decoded_token['nama']

     # Cek apakah user dengan user_id tertentu sudah ada di database
    if user_exists(userid):
        await retry_on_error(update.message.reply_text, text='User tersebut telah terdaftar. Proses tambah user dibatalkan.')

    else:
        # Jika user belum terdaftar, tambahkan user baru ke database
        add_user(userid, nama)
        await retry_on_error(update.message.reply_text, text=f'Berhasil menambahkan user berikut:\nUser ID: {userid}\nNama: {nama}')

    return akhiri_percakapan(context)

@authenticate_admin
async def proses_tambah_admin(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    token = update.message.text
    userid = jwt.get_unverified_header(token)['user_id']
    global jwt_token_keys
    key = jwt_token_keys.pop(userid)

    try:
        decoded_token = jwt.decode(token, key, algorithms=['HS512'])
    except jwt.ExpiredSignatureError:
        await retry_on_error(update.message.reply_text, text='Token telah kadaluarsa. Silahkan membuat token baru dan masukkan kembali token apabila token sudah dibuat kembali oleh calon user admin tersangkut atau batalkan proses tambah user admin dengan menjalankan perintah /batal.')
        return 3

    nama = decoded_token['nama']

    # Cek apakah admin dengan user_id tertentu sudah ada di database
    if user_exists(userid):
        await retry_on_error(update.message.reply_text, text='User tersebut telah terdaftar. Proses tambah admin dibatalkan.')

    else:
        # Jika admin belum terdaftar, tambahkan admin baru ke database
        add_admin(userid, nama)
        await retry_on_error(update.message.reply_text, text=f'Berhasil menambahkan admin berikut:\nUser ID: {userid}\nNama: {nama}')

    return akhiri_percakapan(context)

# Fungsi untuk cek apakah user dengan user_id tertentu sudah ada di database
def user_exists(user_id):
    query = 'SELECT 1 FROM `allowed_users` WHERE `username` = %s'

    global db_connection
    with db_connection.cursor() as cursor:
        cursor.execute(query, (user_id,))
        result = cursor.fetchone()

    return result is not None

# Fungsi untuk menambahkan user baru ke database
def add_user(user_id, nama):
    query = 'INSERT INTO `allowed_users` (`username`, `nama`, `is_admin`) VALUES (%s, %s, 0)'

    global db_connection
    with db_connection.cursor() as cursor:
        cursor.execute(query, (user_id, nama))

    db_connection.commit()

# Fungsi untuk menambahkan admin baru ke database
def add_admin(user_id, nama):
    query = 'INSERT INTO `allowed_users` (`username`, `nama`, `is_admin`) VALUES (%s, %s, 1)'

    global db_connection
    with db_connection.cursor() as cursor:
        cursor.execute(query, (user_id, nama))

    db_connection.commit()

### Hapus User
# Fungsi untuk memperoleh seluruh nama user
def peroleh_data_user():
    query = 'SELECT `username`, `nama` FROM `allowed_users`'

    global db_connection
    with db_connection.cursor() as cursor:
        cursor.execute(query)
        data_user = cursor.fetchall()

    return data_user

# Fungsi untuk menghapus user dari database
def delete_user(username):
    delete_query = 'DELETE FROM `allowed_users` WHERE `username` = %s'

    global db_connection
    with db_connection.cursor() as cursor:
        cursor.execute(delete_query, (user_id,))

    db_connection.commit()

# Fungsi untuk memeriksa apakah user merupakan admin terakhir atau bukan
def is_last_admin(username):
    global db_connection
    cursor = db_connection.cursor()

    # Implementasi pengecekan apakah user merupakan admin atau bukan
    is_admin_query = 'SELECT `is_admin` FROM `allowed_users` WHERE `username` = %s'
    cursor.execute(is_admin_query, (username,))

    if (cursor.fetchone()[0]):
        cursor.close()
        admin_count_query = 'SELECT COUNT(*) FROM `allowed_users` WHERE `is_admin` = 1'
        cursor = db_connection.cursor()
        cursor.execute(admin_count_query)
        admin_count = cursor.fetchone()[0]
    else:
        admin_count = 2

    cursor.close()
    return admin_count <= 1

@authenticate_admin
@check_conv_status
async def hapus_user(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    query = update.callback_query
    await context.bot.edit_message_reply_markup(chat_id=query.message.chat_id, message_id=query.message.message_id, reply_markup=None)
    context.chat_data['in_conversation'] = True

    user_list = peroleh_data_user()
    keyboard = [[InlineKeyboardButton(user[1], callback_data = 'hapus_user_{0}_{1}'.format(user[0], user[1]))] for user in user_list]

    await retry_on_error(context.bot.send_message, chat_id=query.message.chat_id, text='Pilih user yang mau dihapus, atau keluar dari proses dengan menggunakan fungsi /batal.', reply_markup=InlineKeyboardMarkup(keyboard))
    return 1

#  Fungsi untuk memproses hapus user
@authenticate_admin
async def konfirmasi_hapus_user(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    query = update.callback_query

    # Ambil nama user dari inputan chat yang masuk
    username, nama_user = query.data[11:].split('_')

    if is_last_admin(username):
        await retry_on_error(context.bot.send_message, chat_id=query.message.chat_id, text='Tidak bisa menghapus user admin terakhir. Anda dapat menambahkan admin baru terlebih dahulu dan mencoba kembali menghapus user ini. Silahkan pilih kembali user yang mau dihapuskan atau akhiri proses penghapusan user dengan menjalankan perintah /batal.')

    else:
        await context.bot.edit_message_reply_markup(chat_id=query.message.chat_id, message_id=query.message.message_id, reply_markup=None)
        keyboard = [
            [InlineKeyboardButton('Ya', callback_data='konfirmasi_hapus_user_ya')],
            [InlineKeyboardButton('Tidak', callback_data='konfirmasi_hapus_user_tidak')],
        ]
        await retry_on_error(context.bot.send_message, chat_id=query.message.chat_id, text=f'Apakah Anda yakin mau menghapus {nama_user} dari daftar user??', reply_markup=InlineKeyboardMarkup(keyboard))

        # Simpan nama user ke dalam chat data untuk digunakan saat proses hapus user
        context.chat_data['nama_user'] = nama_user
        context.chat_data['username'] = username
        return 2

    return 1

# Fungsi untuk proses hapus user
@authenticate_admin
async def proses_hapus_user(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    query = update.callback_query
    button_pressed = query.data
    await context.bot.edit_message_reply_markup(chat_id=query.message.chat_id, message_id=query.message.message_id, reply_markup=None)

    if button_pressed == 'konfirmasi_hapus_user_ya':
        # Ambil nama user dari chat data
        username = context.chat_data.get('username')
        if not username:
            reply = 'Terjadi kesalahan saat memproses penghapusan user.'
        else:
            # Hapus user dari database
            delete_user(username)
            reply = '{0} berhasil dihapus dari daftar user.'.format(context.chat_data.get('nama_user'))

    elif button_pressed == 'konfirmasi_hapus_user_tidak':
        reply = 'User tidak dihapuskan. Proses penghapusan user dibatalkan.'

    await retry_on_error(context.bot.send_message, chat_id=query.message.chat_id, text=reply)
    return akhiri_percakapan(context)

### Tambah Contact Person
@authenticate_admin
@check_conv_status
async def tambah_cp(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    query = update.callback_query
    await context.bot.edit_message_reply_markup(chat_id=query.message.chat_id, message_id=query.message.message_id, reply_markup=None)
    context.chat_data['in_conversation'] = True

    await retry_on_error(context.bot.send_message, chat_id=query.message.chat_id, text='Silahkan masukkan nomor telepon Telegram contact person.')
    return 1

@authenticate_admin
async def input_tambah_cp(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    nomor = update.message.text

    if nomor[0] == '+':
        if nomor[1:].isnumeric():
            if len(nomor) <= 15 and len(nomor) >= 9:
                if cp_exists(nomor):
                    await retry_on_error(update.message.reply_text, text='Contact person dengan nomor telepon tersebut telah terdaftar dalam basis data. Silahkan masukkan kembali nomor telepon contact person atau akhiri proses dengan menekan tombol /batal')
                    return 1

                context.chat_data['nomor'] = nomor
                await retry_on_error(update.message.reply_text, text='Silahkan masukkan nama contact person.')
            else:
                await retry_on_error(update.message.reply_text, text='Panjang nomor telepon yang dimasukkan tidak sesuai. Silahkan masukkan kembali nomor telepon contact person atau akhiri proses dengan menekan tombol /batal')
                return 1
        else:
            await retry_on_error(update.message.reply_text(text='Nomor telepon yang dimasukkan tidak terdiri dari hanya angka (selain pada kode negara pada karakter pertama). Silahkan masukkan kembali nomor telepon contact person atau akhiri proses dengan menekan tombol /batal'))
            return 1
    else:
        await retry_on_error(update.message.reply_text, text='Nomor telepon yang dimasukkan tidak diawali dengan kode negara dengan karakter \'+\'. Silahkan masukkan kembali nomor telepon contact person atau akhiri proses dengan menekan tombol /batal')
        return 1
    return 2

@authenticate_admin
async def proses_tambah_cp(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    nama = update.message.text
    nomor = context.chat_data.get('nomor')
    add_cp(nomor, nama)
    await retry_on_error(update.message.reply_text, text='Berhasil menambahkan contact person berikut.\nNama: {0}\nNomor HP: {1}'.format(nama, nomor))
    return akhiri_percakapan(context)

# Fungsi untuk cek apakah user dengan user_id tertentu sudah ada di database
def cp_exists(nomor):
    query = 'SELECT 1 FROM `contact_persons` WHERE `phone_num` = %s'

    global db_connection
    with db_connection.cursor() as cursor:
        cursor.execute(query, (nomor,))
        result = cursor.fetchone()

    return result is not None

# Fungsi untuk menambahkan user baru ke database
def add_cp(nama, nomor):
    query = 'INSERT INTO `contact_persons` (`phone_num`, `nama`) VALUES (%s, %s)'

    global db_connection
    with db_connection.cursor() as cursor:
        cursor.execute(query, (nama, nomor))

    db_connection.commit()

### Hapus Contact Person
# Fungsi untuk memperoleh seluruh nama user
def peroleh_data_cp():
    query = 'SELECT `phone_num`, `nama` FROM `contact_persons`'

    global db_connection
    with db_connection.cursor() as cursor:
        cursor.execute(query)
        data_cp = cursor.fetchall()

    return data_cp

# Fungsi untuk menghapus user dari database
def delete_cp(nomor):
    query = 'DELETE FROM `contact_persons` WHERE `phone_num` = %s'

    global db_connection
    with db_connection.cursor() as cursor:
        cursor.execute(query, (nomor,))

    db_connection.commit()

@authenticate_admin
@check_conv_status
async def hapus_cp(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    query = update.callback_query
    await context.bot.edit_message_reply_markup(chat_id=query.message.chat_id, message_id=query.message.message_id, reply_markup=None)
    context.chat_data['in_conversation'] = True

    cp_list = peroleh_data_cp()
    if cp_list == None:
        await retry_on_error(context.bot.send_message, chat_id=query.message.chat_id, text='Belum ada contact person yang terdaftar dalam basis data. Silahkan tambahkan contact person terlebih dahulu. Proses penghapusan contact person diakhiri.')
        return akhiri_percakapan(context)

    keyboard = [[InlineKeyboardButton('{1} ({0})'.format(cp[0], cp[1]), callback_data = 'hapus_cp_{0}_{1}'.format(cp[0], cp[1]))] for cp in cp_list]
    await retry_on_error(context.bot.send_message, chat_id=query.message.chat_id, text='Pilih contact person yang mau dihapus, atau keluar dari proses dengan menggunakan fungsi /batal.', reply_markup=InlineKeyboardMarkup(keyboard))
    return 1

#  Fungsi untuk memproses hapus user
@authenticate_admin
async def konfirmasi_hapus_cp(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    query = update.callback_query

    # Ambil nama user dari inputan chat yang masuk
    nomor, nama = query.data[9:].split('_')
    print(query.data, flush=True)
    print(query.data[9:].split('_'), flush=True)
    print(nama, nomor, flush=True)

    await context.bot.edit_message_reply_markup(chat_id=query.message.chat_id, message_id=query.message.message_id, reply_markup=None)

    keyboard = [
        [InlineKeyboardButton('Ya', callback_data='konfirmasi_hapus_cp_ya')],
        [InlineKeyboardButton('Tidak', callback_data='konfirmasi_hapus_cp_tidak')],
    ]

    await retry_on_error(context.bot.send_message, chat_id=query.message.chat_id, text=f'Apakah Anda yakin mau menghapus {nama} ({nomor}) dari daftar contact person?', reply_markup=InlineKeyboardMarkup(keyboard))

    # Simpan nama user ke dalam chat data untuk digunakan saat proses hapus user
    context.chat_data['nama_cp'] = nama
    context.chat_data['nomor'] = nomor

    return 2

# Fungsi untuk proses hapus user
@authenticate_admin
async def proses_hapus_cp(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    query = update.callback_query
    button_pressed = query.data
    await context.bot.edit_message_reply_markup(chat_id=query.message.chat_id, message_id=query.message.message_id, reply_markup=None)

    if button_pressed == 'konfirmasi_hapus_cp_ya':
        # Ambil nama user dari chat data
        nomor = context.chat_data.get('nomor')
        if not nomor:
            reply = 'Terjadi kesalahan saat memproses penghapusan contact person.'
        else:
            # Hapus cp dari basis data
            delete_cp(nomor)
            text = '{0} ({1}) berhasil dihapus dari daftar contact person.'.format(context.chat_data.get('nama_cp'), nomor)

        await retry_on_error(context.bot.send_message, chat_id=query.message.chat_id, text=text)
    elif button_pressed == 'konfirmasi_hapus_cp_tidak':
        await retry_on_error(context.bot.send_message, chat_id=query.message.chat_id, text='Contact person tidak dihapuskan. Proses penghapusan contact person dibatalkan.')

    return akhiri_percakapan(context)

### Peroleh Lokasi
@authenticate_user
@check_conv_status
async def peroleh_lokasi(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    query = update.callback_query
    await context.bot.edit_message_reply_markup(chat_id=query.message.chat_id, message_id=query.message.message_id, reply_markup=None)
    context.chat_data['in_conversation'] = True
    site_list = peroleh_nama_site()
    context.chat_data['site_list'] = site_list
    keyboard = []

    for i in range(ceil(len(site_list)/3)):
        temp_item = site_list[i * 3]
        keyboard.append([InlineKeyboardButton(temp_item, callback_data='item_' + temp_item)])

        for j in range(1, 3):
            try:
                temp_item = site_list[i * 3 + j]
                keyboard[-1].append(InlineKeyboardButton(temp_item, callback_data='item_' + temp_item))
            except IndexError:
                break

    message = await retry_on_error(context.bot.send_message, chat_id=query.message.chat_id, text='Silahkan masukkan nama item atau pilih item dari tombol di bawah ini.', reply_markup=InlineKeyboardMarkup(keyboard))
    context.user_data['last_message_id'] = message.message_id
    return 1

# Fungsi untuk memproses peroleh lokasi
@authenticate_user
async def proses_peroleh_lokasi_text(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    await context.bot.edit_message_reply_markup(chat_id=update.effective_chat.id, message_id=context.user_data.get('last_message_id'), reply_markup=None)
    text = update.message.text
    if len(text) == 8 and text.isalnum():

        # Cek apakah site dengan Site_ID_Tenant tertentu ada di database
        try:
            site = peroleh_data_site(next(item for item in context.chat_data['site_list'] if item == text.upper()))
            await kirim_data_item(update, context, [False, False], site)
            return akhiri_percakapan(context)
        except StopIteration:
            await retry_on_error(update.message.reply_text, text='Site tidak ditemukan. Silahkan masukkan atau pilih kembali nama site atau keluar dari proses dengan menggunakan fungsi /batal.')
            return 1

    await retry_on_error(update.message.reply_text, text='Format penamaan salah. Silahkan masukkan atau pilih kembali nama site atau keluar dari proses dengan menggunakan fungsi /batal.')
    return 1

@authenticate_user
async def proses_peroleh_lokasi_button(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    query = update.callback_query
    await context.bot.edit_message_reply_markup(chat_id=query.message.chat_id, message_id=query.message.message_id, reply_markup=None)

    # Cek apakah site dengan Site_ID_Tenant tertentu ada di database
    try:
        site = peroleh_data_site(next(item for item in context.chat_data['site_list'] if item == query.data[5:]))
        await kirim_data_item(update, context, [False, False], site)
        return akhiri_percakapan(context)
    except StopIteration:
        await retry_on_error(context.bot.send_message, chat_id=query.message.chat_id, text='Site tidak ditemukan. Silahkan masukkan atau pilih kembali nama site atau keluar dari proses dengan menggunakan fungsi /batal.')
        return 1

# Fungsi untuk memparse koordinat
def parse_coordinates(coordinates):
    coordinates = coordinates.split(',')
    latitude = float(''.join(filter(lambda c: c.isdigit() or c in ['.', '-'], coordinates[0].strip())))
    longitude = float(''.join(filter(lambda c: c.isdigit() or c in ['.', '-'], coordinates[1].strip())))
    return latitude, longitude


@authenticate_user
@check_conv_status
async def peroleh_lokasi_func(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    args = context.args
    if args:
        argument = args[0]
        if len(argument) == 8 and argument.isalnum():

            # Cek apakah site dengan Site_ID_Tenant tertentu ada di database
            site = peroleh_data_site(argument.upper())
            print(site, flush = True)

            if site:
                await kirim_data_item(update, context, [False, False], site)

            else:
                await retry_on_error(update.message.reply_text, text='Site tidak ditemukan.')

        else:
            await retry_on_error(update.message.reply_text, text='Format penamaan salah.')
    else:
        await retry_on_error(update.message.reply_text, text='Sintaks: /peroleh_lokasi [nama item]\nContoh: /peroleh_lokasi 20BAT001')

### Peroleh Nama
@authenticate_user
@check_conv_status
async def peroleh_nama_1(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    query = update.callback_query
    context.chat_data['in_conversation'] = True
    await context.bot.edit_message_reply_markup(chat_id=query.message.chat_id, message_id=query.message.message_id, reply_markup=None)
    await retry_on_error(context.bot.send_message, chat_id=query.message.chat_id, text='Kirimkan Lokasi Pencarian.')
    return 1

@authenticate_user
async def peroleh_nama_2(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    query = update.callback_query
    location = update.message.location
    context.chat_data['latitude'] = location.latitude
    context.chat_data['longitude'] = location.longitude
    await retry_on_error(update.message.reply_text, text='Kirimkan radius pencarian dalam satuan meter.')
    return 2

# Fungsi async untuk memproses input pengguna
async def proses_peroleh_nama(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    dist = update.message.text

    if dist.lstrip('-').replace(',', '').replace('.', '').isnumeric():
        radius_meter = float(dist.replace(',', '.'))

        if radius_meter <= 0:
            await retry_on_error(update.message.reply_text, text='Nilai radius harus lebih besar dari 0. Silahkan masukkan kembali nilai radius atau keluar dari proses dengan menggunakan fungsi /batal.')
            return 2

        else:
            # Dapatkan lokasi pengguna dari chat_data
            user_latitude = float(context.chat_data['latitude'])
            user_longitude = float(context.chat_data['longitude'])

            # Hitung daftar situs dalam radius
            sites_in_radius = peroleh_dari_radius(user_latitude, user_longitude, radius_meter)

            if sites_in_radius:
                for site in sites_in_radius:
                    await kirim_data_item(update, context, [False, False], site)
                return akhiri_percakapan(context)
            else:
                await retry_on_error(update.message.reply_text, text=f'Tidak ada item dalam radius {radius_meter} meter dari lokasi. Silahkan masukkan kembali nilai radius atau keluar dari proses dengan menggunakan fungsi /batal.')
                return 2

    await retry_on_error(update.message.reply_text, text='Radius harus berupa bilangan cacah atau pecahan desimal > 0 dalam satuan meter. Contoh: 1000. Silahkan masukkan kembali nilai radius atau keluar dari proses dengan menggunakan fungsi /batal.')
    return 2

@authenticate_user
@check_conv_status
async def peroleh_berkas(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    query = update.callback_query
    context.chat_data['in_conversation'] = True
    await context.bot.edit_message_reply_markup(chat_id=query.message.chat_id, message_id=query.message.message_id, reply_markup=None)
    await retry_on_error(context.bot.send_message, chat_id=query.message.chat_id, text='Masukkan nama site.')
    return 1

@authenticate_user
async def proses_peroleh_berkas(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    site_name = update.message.text
    file_path = f'/files/{site_name}.txt'

    if os.path.exists(file_path):
        chat_id = update.effective_chat.id
        await retry_on_error(context.bot.send_document, chat_id=chat_id, document=open(file_path, 'rb'))
    else:
        await retry_on_error(update.message.reply_text, text="Berkas tidak ditemukan. Silahkan masukkan kembali nama berkas atau keluar dari proses dengan menggunakan fungsi /batal.")
        return 1

    return akhiri_percakapan(context)

async def batal(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    if context.chat_data['in_conversation']:
        try:
            await retry_on_error(update.message.reply_text, text='Proses telah dibatalkan.')
        except AttributeError:
            await retry_on_error(context.bot.send_message, chat_id=query.message.chat_id, text='Proses telah dibatalkan.', reply_markup=InlineKeyboardMarkup(keyboard))
        finally:
            return akhiri_percakapan(context)

    await retry_on_error(update.message.reply_text, text='Tidak ada proses yang sedang berjalan.')

def akhiri_percakapan(context: ContextTypes.DEFAULT_TYPE):
    context.chat_data['in_conversation'] = False
    return ConversationHandler.END

@authenticate_user
async def kirim_bantuan(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    query = update.callback_query
    try:
        await context.bot.edit_message_reply_markup(chat_id=query.message.chat_id, message_id=query.message.message_id, reply_markup=None)
    except Exception:
        pass
    text='''Fungsi Dasar
/start
Memulai percakapan. Percakapan tidak dapat dimulai apabila sebuah proses sedang berjalan. Lihat /batal.
/batal
Mengakhiri proses. Mengakhiri proses yang sedang berjalan dan membuka kemungkinan untuk memulai kembali percakapan.
/bantuan
Melihat bantuan penggunaan (teks ini).
/peroleh\_token
Peroleh token untuk keperluan autentikasi.
/peroleh\_username
Peroleh username untuk keperluan inisialisasi.

Arahan Menu
Menu Utama: Berisi berbagai pilihan untuk membuka submenu.
Input Data: Menjalankan proses untuk memasukkan data ke dalam basis data.
Peroleh Lokasi Item: Menjalankan proses untuk memperoleh lokasi item berdasarkan nama item.
Peroleh Nama Item: Menjalankan proses untuk memperoleh nama-nama item berdasarkan lokasi.
Peroleh File Site: Menjalankan proses untuk memperoleh berkas-berkas berdasarkan nama.'''
    cp_list = peroleh_cp()
    if cp_list:
        text += '\n\nApabila memerlukan bantuan tambahan, anda dapat menghubungi helpdesk pada kontak berikut:\n'
        text += '\n'.join(['{0}: [{1}](https://t.me/{1})'.format(cp[0], cp[1]) for cp in peroleh_cp()])

    try:
        await retry_on_error(context.bot.send_message, chat_id=query.message.chat_id, text=text, parse_mode=ParseMode.MARKDOWN)
    except Exception as e:
        print(type(e), e, flush = True)
        await retry_on_error(update.message.reply_text, text=text, parse_mode=ParseMode.MARKDOWN)

async def peroleh_token(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    await retry_on_error(update.message.reply_text, text=f'Mohon kirimkan nama lengkap anda.')
    return 1

async def peroleh_token_process(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    nama = update.message.text.capitalize()
    userid = str(update.message.from_user.id)
    expiration_time = datetime.utcnow() + timedelta(minutes=5)

    global jwt_token_keys
    jwt_token_keys[userid] = secrets.token_bytes(32)

    token = jwt.encode(
        {
            'nama': nama,
            'exp': expiration_time
        },
        jwt_token_keys[userid],
        algorithm = 'HS512',
        headers = {'user_id': userid}
    )
    await retry_on_error(update.message.reply_text, text=f'Token anda adalah: {token}')
    return akhiri_percakapan(context)

async def peroleh_username(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status=[False, False]):
    await retry_on_error(update.message.reply_text, text=f'Username anda adalah: {str(update.message.from_user.id)}')

@authenticate_user
async def kirim_data_item(update: Update, context: ContextTypes.DEFAULT_TYPE, auth_status, data):
    text = 'Site\_ID\_Tenant: {}\n'.format(data[0])
    text += 'Tenant: {}\n'.format(data[1])
    text += 'Alamat: {}\n'.format(data[2]).translate(str.maketrans({'\\': '\\\\',
                                                                    '#': r'\#',
                                                                    '*': r'\*',
                                                                    '_': r'\_',
                                                                    '{': r'\{',
                                                                    '}': r'\}',
                                                                    '(': r'\(',
                                                                    ')': r'\)',
                                                                    '<': r'\<',
                                                                    '>': r'\>',
                                                                    '[': r'\[',
                                                                    ']': r'\]',
                                                                    '|': r'\|',
                                                                    '`': r'\`',
                                                                    '!': r'\!',
                                                                    ':': r'\:',
                                                                    '=': r'\=',
                                                                    '~': r'\~',
                                                                    '^': r'\^'}))
    text += 'Koordinat Site: [{0},{1}](http://maps.google.com/maps?q={0},{1})\n'.format(data[3], data[4])
    query = update.callback_query

    if query:
        await retry_on_error(context.bot.send_message, chat_id=query.message.chat_id, text=text, parse_mode=ParseMode.MARKDOWN)
        try:
            await retry_on_error(context.bot.send_location, chat_id=query.message.chat_id, latitude=data[3], longitude=data[4])
        except Exception as e:
            print(e, flush = True)
    else:
        await retry_on_error(update.message.reply_text, text=text, parse_mode=ParseMode.MARKDOWN)
        try:
            await retry_on_error(context.bot.send_location, chat_id=query.message.chat_id, latitude=data[3], longitude=data[4])
        except Exception as e:
            print(e, flush = True)

def peroleh_cp():
    query = 'SELECT `nama`, `phone_num` FROM `contact_persons`'

    global db_connection
    with db_connection.cursor() as cursor:
        cursor.execute(query)
        contacts = cursor.fetchall()
    return contacts

##### MAIN
if __name__ == '__main__':
    main()
